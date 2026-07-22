"""Weekly SAM.gov construction contract award collector.

Pulls awards signed in the previous Monday-Sunday week for the Construction
NAICS family, appends them to a running Excel workbook, and optionally
emails a success/failure notification. Run manually or from a scheduler
(Windows Task Scheduler / cron) with no arguments for a normal weekly run.

Data comes from sam.gov's own internal search UI API (sgs/v1/search +
opps/v2/opportunities), driven through a real headless Chromium browser via
Playwright. SAM.gov's public api.sam.gov REST endpoint proved unreachable
with a standard public API key during development; this mirrors the exact
browser workflow used to pull this data manually. Being an undocumented
internal API, it could change without notice - if this script starts
failing, that's the first thing to check.
"""

import argparse
import html
import logging
import os
import re
import smtplib
import sys
from datetime import date, datetime, time, timedelta
from email.message import EmailMessage
from pathlib import Path
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import sync_playwright
from tenacity import retry, retry_if_exception_type, stop_after_attempt, wait_exponential

SAM_ORIGIN = "https://sam.gov"
SEARCH_URL = "https://sam.gov/api/prod/sgs/v1/search/"
DETAIL_URL_TEMPLATE = "https://sam.gov/api/prod/opps/v2/opportunities/{notice_id}"
DEFAULT_NAICS_CODES = "23,237,2362,2373,2379"
BROWSER_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
EASTERN = ZoneInfo("America/New_York")

EXCEL_COLUMNS = [
    "Awardee",
    "Title",
    "Owner",
    "Project Value",
    "Contract Award Date",
    "Contract Award Number",
    "Description",
]
COLUMN_WIDTHS = {
    "Awardee": 30,
    "Title": 35,
    "Owner": 30,
    "Project Value": 16,
    "Contract Award Date": 18,
    "Contract Award Number": 22,
    "Description": 60,
}

_HTML_TAG_RE = re.compile(r"<[^>]+>")


class SamGovRateLimited(Exception):
    pass


class SamGovServerError(Exception):
    pass


# ---------------------------------------------------------------------------
# Config / logging
# ---------------------------------------------------------------------------

def _env_bool(name, default=False):
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in ("1", "true", "yes", "on")


def _env_int(name, default):
    value = os.environ.get(name)
    if value is None or not value.strip():
        return default
    try:
        return int(value)
    except ValueError:
        return default


def load_config():
    return {
        "naics_codes": os.environ.get("NAICS_CODES", DEFAULT_NAICS_CODES).strip(),
        "modified_date_lookback_days": _env_int("MODIFIED_DATE_LOOKBACK_DAYS", 30),
        "cdp_url": os.environ.get("CDP_URL", "http://localhost:9222"),
        "excel_output_path": os.environ.get("EXCEL_OUTPUT_PATH", "./sam_construction_awards.xlsx"),
        "email_enabled": _env_bool("EMAIL_ENABLED", False),
        "smtp_host": os.environ.get("SMTP_HOST", "smtp.gmail.com"),
        "smtp_port": _env_int("SMTP_PORT", 465),
        "smtp_use_ssl": _env_bool("SMTP_USE_SSL", True),
        "email_address": os.environ.get("EMAIL_ADDRESS", ""),
        "email_password": os.environ.get("EMAIL_PASSWORD", ""),
        "notify_email_to": os.environ.get("NOTIFY_EMAIL_TO", ""),
        "log_level": os.environ.get("LOG_LEVEL", "INFO"),
        "log_path": os.environ.get("LOG_PATH", "./sam_automation.log"),
    }


def setup_logging(log_path, log_level):
    level = getattr(logging, str(log_level).upper(), logging.INFO)
    root = logging.getLogger()
    root.setLevel(level)
    root.handlers.clear()

    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    root.addHandler(file_handler)

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    root.addHandler(console_handler)


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def get_previous_week_range(reference_date=None):
    """Return (monday, sunday) of the full week before the one containing reference_date."""
    reference_date = reference_date or date.today()
    this_monday = reference_date - timedelta(days=reference_date.weekday())
    previous_monday = this_monday - timedelta(days=7)
    previous_sunday = previous_monday + timedelta(days=6)
    return previous_monday, previous_sunday


def _parse_iso_date(value):
    if not value:
        raise ValueError("empty date value")
    return date.fromisoformat(str(value)[:10])


def _sam_date_param(d):
    """sam.gov's search UI sends modified_date.from/to as YYYY-MM-DD plus the Eastern
    Time UTC offset (e.g. 2026-07-08-04:00) - computed here rather than hardcoded so it
    stays correct across the EST/EDT transition."""
    dt = datetime.combine(d, time.min, tzinfo=EASTERN)
    offset_minutes = int(dt.utcoffset().total_seconds() // 60)
    sign = "-" if offset_minutes < 0 else "+"
    hours, minutes = divmod(abs(offset_minutes), 60)
    return f"{d.isoformat()}{sign}{hours:02d}:{minutes:02d}"


# ---------------------------------------------------------------------------
# sam.gov session + HTTP helpers
# ---------------------------------------------------------------------------

def bootstrap_session(playwright, cdp_url):
    """Connects to an already-running Chrome window (started with
    --remote-debugging-port and logged into/browsed to sam.gov/search by a human) rather
    than launching a fresh automated browser. sam.gov's session-setup call never fires for
    a Playwright-launched browser - this appears to be automation detection - so this
    reuses a genuinely human-initiated session instead of trying to spoof around that."""
    try:
        browser = playwright.chromium.connect_over_cdp(cdp_url)
    except Exception as exc:
        raise RuntimeError(
            f"Could not connect to a Chrome window at {cdp_url}. Make sure Chrome is running "
            "with --remote-debugging-port and you've navigated it to https://sam.gov/search. "
            f"Underlying error: {exc}"
        ) from exc

    if not browser.contexts:
        raise RuntimeError(f"Connected to Chrome at {cdp_url}, but it has no open browser context/tabs.")
    context = browser.contexts[0]

    page_urls = [p.url for p in context.pages]
    all_cookies = context.cookies()
    logging.debug("Connected Chrome pages: %s", page_urls)
    logging.debug("Connected Chrome cookies: %s", [c["name"] for c in all_cookies])

    token = None
    for cookie in all_cookies:
        if cookie["name"] in ("XSRF-TOKEN", "SESSION"):
            token = cookie["value"]
            break
    if not token:
        raise RuntimeError(
            "No sam.gov session cookie found in the connected Chrome window. Make sure "
            "you've navigated that window to https://sam.gov/search and let it fully load. "
            f"Open page(s): {page_urls}. Cookies seen: {[c['name'] for c in all_cookies]}"
        )
    return browser, context, token


@retry(
    retry=retry_if_exception_type((PlaywrightError, SamGovRateLimited, SamGovServerError)),
    wait=wait_exponential(multiplier=2, min=2, max=60),
    stop=stop_after_attempt(5),
    reraise=True,
)
def _pw_get_json(api_context, url, params, token):
    headers = {
        "accept": "application/json, text/plain, */*",
        "x-auth-token": token,
        "x-xsrf-token": token,
    }
    response = api_context.get(url, params=params, headers=headers, timeout=30000)
    if response.status == 429:
        raise SamGovRateLimited(f"Rate limited (429) calling {url}")
    if response.status >= 500:
        raise SamGovServerError(f"Server error ({response.status}) calling {url}")
    if not response.ok:
        raise RuntimeError(f"{response.status} error calling {url}: {response.text()[:500]}")
    return response.json()


# ---------------------------------------------------------------------------
# Fetching
# ---------------------------------------------------------------------------

def fetch_bulk_records(api_context, token, start_date, end_date, naics_codes, lookback_days):
    """sam.gov's search index only filters by modified_date (when the notice was last
    touched), not the real award date, so this queries a padded window; results are
    narrowed to the target award week later using the per-record detail lookup."""
    lookback_start = start_date - timedelta(days=lookback_days)
    records = []
    page_num = 0
    page_size = 100
    while True:
        params = {
            "index": "opp",
            "page": page_num,
            "sort": "-modifiedDate",
            "size": page_size,
            "mode": "search",
            "responseType": "json",
            "q": "",
            "qMode": "ALL",
            "modified_date.to": _sam_date_param(end_date),
            "modified_date.from": _sam_date_param(lookback_start),
            "naics": naics_codes,
            "notice_type": "a",
        }
        data = _pw_get_json(api_context, SEARCH_URL, params, token)
        batch = data.get("_embedded", {}).get("results", [])
        records.extend(batch)
        total = data.get("page", {}).get("totalElements", len(records))
        page_num += 1
        if not batch or page_num * page_size >= total:
            break
    return records


def fetch_award_detail(api_context, token, notice_id):
    try:
        return _pw_get_json(api_context, DETAIL_URL_TEMPLATE.format(notice_id=notice_id), {"api_key": "null"}, token)
    except Exception as exc:
        logging.warning("Could not fetch award detail for notice %s: %s", notice_id, exc)
        return None


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_bulk_record(record):
    org_hierarchy = record.get("organizationHierarchy") or []
    owner = org_hierarchy[0].get("name", "") if org_hierarchy else ""
    descriptions = record.get("descriptions") or []
    description_html = descriptions[0].get("content", "") if descriptions else ""
    return {
        "notice_id": record.get("_id", ""),
        "Awardee": ((record.get("award") or {}).get("awardee") or {}).get("name", ""),
        "Title": record.get("title", ""),
        "Owner": owner,
        "Description": html.unescape(_HTML_TAG_RE.sub(" ", description_html)).strip(),
    }


def merge_award_detail(parsed_record, detail_data):
    """Fills in the fields only present on the per-contract detail page: the real
    Contract Award Date, Contract Award Number, and Project Value."""
    award = ((detail_data or {}).get("data2") or {}).get("award") or {}
    parsed_record["Contract Award Date"] = award.get("date", "")
    parsed_record["Contract Award Number"] = award.get("number", "")
    parsed_record["Project Value"] = award.get("amount", "")
    awardee_name = (award.get("awardee") or {}).get("name")
    if awardee_name:
        parsed_record["Awardee"] = awardee_name
    return parsed_record


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def format_workbook(worksheet):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, header in enumerate(EXCEL_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        worksheet.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 20)
    worksheet.freeze_panes = "A2"


def load_or_create_workbook(path):
    if path.exists():
        return load_workbook(path)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Awards"
    worksheet.append(EXCEL_COLUMNS)
    format_workbook(worksheet)
    workbook.save(path)
    return workbook


def append_records(workbook, records, path):
    worksheet = workbook["Awards"] if "Awards" in workbook.sheetnames else workbook.active
    for record in records:
        worksheet.append([record.get(column, "") for column in EXCEL_COLUMNS])
    try:
        workbook.save(path)
    except PermissionError:
        logging.error("Could not save %s - is it open in Excel? Close the file and re-run.", path)
        raise
    return len(records)


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

def send_email(config, subject, body):
    if not config["email_enabled"]:
        logging.info("Email notifications disabled (EMAIL_ENABLED=false); skipping send.")
        return
    if not config["email_address"] or not config["notify_email_to"]:
        logging.warning("EMAIL_ENABLED is true but EMAIL_ADDRESS/NOTIFY_EMAIL_TO are not set; skipping send.")
        return
    try:
        message = EmailMessage()
        message["Subject"] = subject
        message["From"] = config["email_address"]
        message["To"] = config["notify_email_to"]
        message.set_content(body)

        if config["smtp_use_ssl"]:
            with smtplib.SMTP_SSL(config["smtp_host"], config["smtp_port"], timeout=30) as smtp:
                smtp.login(config["email_address"], config["email_password"])
                smtp.send_message(message)
        else:
            with smtplib.SMTP(config["smtp_host"], config["smtp_port"], timeout=30) as smtp:
                smtp.starttls()
                smtp.login(config["email_address"], config["email_password"])
                smtp.send_message(message)
        logging.info("Notification email sent to %s", config["notify_email_to"])
    except Exception:
        logging.exception("Failed to send notification email")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(description="Weekly SAM.gov construction award collector")
    parser.add_argument(
        "--reference-date",
        help="ISO date (YYYY-MM-DD) to compute the target week from; defaults to today. Useful for backfills/testing.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Fetch and parse records but skip writing to Excel and sending email.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    load_dotenv()
    config = load_config()
    setup_logging(config["log_path"], config["log_level"])

    try:
        reference_date = date.fromisoformat(args.reference_date) if args.reference_date else None
        start_date, end_date = get_previous_week_range(reference_date)
        logging.info("Target award week: %s to %s", start_date, end_date)

        with sync_playwright() as playwright:
            # bootstrap_session connects to a Chrome window you already have open and
            # logged into sam.gov with - it's not ours to close, so no browser.close() here.
            _browser, context, token = bootstrap_session(playwright, config["cdp_url"])
            api_context = context.request
            bulk_records = fetch_bulk_records(
                api_context,
                token,
                start_date,
                end_date,
                config["naics_codes"],
                config["modified_date_lookback_days"],
            )
            logging.info("Fetched %d candidate record(s) from sam.gov search", len(bulk_records))

            parsed_records = []
            for record in bulk_records:
                parsed = parse_bulk_record(record)
                notice_id = parsed.pop("notice_id", "")
                if not notice_id:
                    continue
                detail_data = fetch_award_detail(api_context, token, notice_id)
                parsed = merge_award_detail(parsed, detail_data)
                try:
                    award_date = _parse_iso_date(parsed.get("Contract Award Date"))
                except ValueError:
                    logging.debug("Skipping notice %s with missing/unparsable award date", notice_id)
                    continue
                if not (start_date <= award_date <= end_date):
                    continue
                parsed_records.append(parsed)
            logging.info("%d record(s) fall within the target award week", len(parsed_records))

        if args.dry_run:
            logging.info(
                "[DRY RUN] Would append %d record(s) to %s; skipping Excel write and email.",
                len(parsed_records),
                config["excel_output_path"],
            )
            return

        excel_path = Path(config["excel_output_path"])
        workbook = load_or_create_workbook(excel_path)
        added = append_records(workbook, parsed_records, excel_path)
        logging.info("Appended %d row(s) to %s", added, excel_path)

        subject = f"SAM.gov Award Automation: {added} construction award(s) added ({start_date} to {end_date})"
        body = (
            f"SAM.gov construction award collection completed successfully.\n\n"
            f"Week: {start_date} to {end_date}\n"
            f"Rows added: {added}\n"
            f"Excel file: {excel_path.resolve()}\n"
            f"Log file: {Path(config['log_path']).resolve()}\n"
        )
        send_email(config, subject, body)
        logging.info("Run completed successfully.")
    except Exception as exc:
        logging.exception("Run failed")
        subject = "SAM.gov Award Automation: FAILED"
        body = (
            f"The weekly SAM.gov automation run failed.\n\n"
            f"Error: {exc}\n\n"
            f"See log file for full details: {Path(config['log_path']).resolve()}\n"
        )
        try:
            send_email(config, subject, body)
        except Exception:
            logging.exception("Additionally failed to send the failure notification email")
        sys.exit(1)


if __name__ == "__main__":
    main()
